import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import streamlit as st
from sklearn.compose import ColumnTransformer
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.impute import SimpleImputer
from sklearn.metrics import classification_report, roc_auc_score, roc_curve, confusion_matrix
from sklearn.model_selection import train_test_split
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import OneHotEncoder

# ── PAGE CONFIG ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="ICO Data Security Incident Explorer",
    page_icon="🔐",
    layout="wide",
    initial_sidebar_state="expanded",
)

TPL        = "plotly_white"
C_CYBER    = "#dc2626"
C_NONCYBER = "#2563eb"
C_NEUTRAL  = "#0f766e"
C_WARN     = "#f59e0b"
C_PURPLE   = "#7c3aed"

# ── CUSTOM CSS ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
.kpi-card{background:linear-gradient(135deg,#1e293b 0%,#0f172a 100%);border:1px solid #334155;border-radius:12px;padding:18px 20px 14px;text-align:center;height:110px;display:flex;flex-direction:column;justify-content:center;}
.kpi-label{font-size:0.70rem;text-transform:uppercase;letter-spacing:.08em;color:#94a3b8;margin-bottom:6px;}
.kpi-value{font-size:1.65rem;font-weight:800;color:#f1f5f9;line-height:1.1;}
.kpi-sub{font-size:0.68rem;color:#64748b;margin-top:4px;}
.hero{background:linear-gradient(135deg,#0f172a 0%,#1e3a5f 50%,#0f172a 100%);border:1px solid #1e40af;border-radius:16px;padding:28px 32px;margin-bottom:24px;}
.hero h1{color:#f1f5f9;font-size:1.9rem;font-weight:800;margin:0 0 8px;}
.hero p{color:#94a3b8;font-size:0.95rem;margin:0;line-height:1.6;}
.section-tag{display:inline-block;background:#1e3a5f;color:#93c5fd;border-radius:6px;padding:2px 10px;font-size:0.72rem;font-weight:600;letter-spacing:.06em;text-transform:uppercase;margin-bottom:8px;}
button[data-baseweb="tab"]{font-size:0.82rem;}
section[data-testid="stSidebar"]{background:#0f172a;}
section[data-testid="stSidebar"] *{color:#e2e8f0 !important;}
section[data-testid="stSidebar"] .stMultiSelect [data-baseweb="tag"]{background:#1e40af !important;}
</style>
""", unsafe_allow_html=True)

# ── HELPERS ────────────────────────────────────────────────────────────────────
def note(title, body):
    with st.expander("ℹ️  " + title, expanded=False):
        st.markdown(body)

def kpi(label, value, sub=""):
    return ('<div class="kpi-card"><div class="kpi-label">'+label+'</div>'
            '<div class="kpi-value">'+str(value)+'</div>'
            '<div class="kpi-sub">'+sub+'</div></div>')

# ── EXCEL EXPORT FUNCTION ──────────────────────────────────────────────────────
def build_excel_export(df_export):
    from openpyxl import Workbook as _WB
    from openpyxl.styles import Font as _F, PatternFill as _PF, Alignment as _A, Border as _B, Side as _S
    from openpyxl.utils import get_column_letter as _gcl
    from openpyxl.formatting.rule import ColorScaleRule as _CSR, DataBarRule as _DBR
    from openpyxl.worksheet.table import Table as _T, TableStyleInfo as _TSI
    from openpyxl.worksheet.hyperlink import Hyperlink as _H
    from io import BytesIO as _IO
    from datetime import datetime as _dt

    NAVY="0F2540";TEAL="0E7C7B";SLATE="2C3E50";PURPLE="6C3483"
    AMBER="D68910";WHITE="FFFFFF";LBG="F4F6F8";CYAN="D6EAF8"
    GREEN="D5F5E3";RED="FADBD8";AMB="FDEBD0";PUR="E8DAEF";GBD="BDC3C7"

    def _fnt(bold=False,size=11,color="000000",italic=False):
        return _F(name="Calibri",bold=bold,size=size,color=color,italic=italic)
    def _fl(c): return _PF("solid",fgColor=c)
    def _ctr(): return _A(horizontal="center",vertical="center",wrap_text=True)
    def _lft(): return _A(horizontal="left",vertical="center",wrap_text=True,indent=1)
    def _rgt(): return _A(horizontal="right",vertical="center")
    def _thin():
        s=_S(style="thin",color=GBD); return _B(left=s,right=s,top=s,bottom=s)
    def _thickb():
        t=_S(style="medium",color=NAVY); n=_S(style="thin",color=GBD)
        return _B(left=n,right=n,top=n,bottom=t)
    def _hdr(cell,bg=NAVY,fg=WHITE,sz=11):
        cell.font=_fnt(bold=True,size=sz,color=fg);cell.fill=_fl(bg)
        cell.alignment=_ctr();cell.border=_thickb()
    def _sec(ws,row,c1,c2,text,bg=TEAL,fg=WHITE,sz=11):
        ws.merge_cells(start_row=row,start_column=c1,end_row=row,end_column=c2)
        cell=ws.cell(row=row,column=c1,value=text)
        cell.font=_fnt(bold=True,size=sz,color=fg);cell.fill=_fl(bg)
        cell.alignment=_A(horizontal="left",vertical="center",indent=2)
        ws.row_dimensions[row].height=26
    def _aw(ws,ci,mn=10,mx=38,ex=2):
        ml=0
        for row in ws.iter_rows(min_col=ci,max_col=ci):
            for c in row:
                if c.value: ml=max(ml,len(str(c.value)))
        ws.column_dimensions[_gcl(ci)].width=min(max(ml+ex,mn),mx)

    RAW_C=['BI_Reference','Year','Quarter','Sector','Data_Subject_Type','Data_Type',
           'Incident_Category','Incident_Type','Decision_Taken',
           'No_Data_Subjects_Affected','Time_Taken_to_Report']
    ENG_C=['Is_Cyber','Is_High_Impact','Impact_Score','Is_Special_Category',
           'Within_72hrs','Severity_Score','Sector_Risk_Tier']
    ALL_C=RAW_C+ENG_C
    CDN={'BI_Reference':'Breach ID','Year':'Year','Quarter':'Quarter','Sector':'Sector',
         'Data_Subject_Type':'Data Subjects','Data_Type':'Data Category',
         'Incident_Category':'Breach Category','Incident_Type':'Breach Type',
         'Decision_Taken':'ICO Decision','No_Data_Subjects_Affected':'People Affected',
         'Time_Taken_to_Report':'Reporting Time','Is_Cyber':'Is Cyber? (0/1)',
         'Is_High_Impact':'High Impact? (0/1)','Impact_Score':'Impact Score (1-6)',
         'Is_Special_Category':'Special Category? (0/1)','Within_72hrs':'Within 72hrs? (0/1)',
         'Severity_Score':'Severity Score (0-11)','Sector_Risk_Tier':'Sector Risk Tier'}

    wb=_WB()

    # Sheet 1: Overview
    ws0=wb.active;ws0.title="Overview";ws0.sheet_view.showGridLines=False
    for col,w in [("A",3),("B",38),("C",22),("D",50)]: ws0.column_dimensions[col].width=w
    ws0.merge_cells("B2:D2")
    c=ws0["B2"];c.value="ICO Data Security Incident Explorer"
    c.font=_F(name="Calibri",bold=True,size=20,color=WHITE);c.fill=_fl(NAVY)
    c.alignment=_A(horizontal="left",vertical="center",indent=2);ws0.row_dimensions[2].height=48
    ws0.merge_cells("B3:D3")
    c2=ws0["B3"];c2.value="Personal data breach analysis  |  UK ICO  |  2019 - Q4 2025"
    c2.font=_F(name="Calibri",size=11,color="AABCD0",italic=True);c2.fill=_fl(NAVY)
    c2.alignment=_A(horizontal="left",vertical="center",indent=2);ws0.row_dimensions[3].height=22
    ws0.row_dimensions[4].height=10
    _sec(ws0,5,2,4,"  WORKBOOK CONTENTS",NAVY,WHITE,12)
    for ci,h in enumerate(["Sheet","Contents","Description"],2):
        _hdr(ws0.cell(6,ci,h),bg=TEAL);ws0.row_dimensions[6].height=22
    idx=[("Data","Filtered breach data","Raw + engineered columns, Excel Table, conditional formatting"),
         ("Dictionary","Field definitions","All 18 columns with ICO / UK GDPR definitions"),
         ("Methods","Cleaning and engineering","Step-by-step methods with Python code"),
         ("Stats","Summary statistics","KPIs, year breakdown, top sectors")]
    for ri,(sh,ct,desc) in enumerate(idx,7):
        bg=LBG if ri%2==0 else WHITE
        for ci,val in enumerate([sh,ct,desc],2):
            c=ws0.cell(ri,ci,val);c.fill=_fl(bg);c.font=_fnt(size=10)
            c.alignment=_lft();c.border=_thin()
        lnk=ws0.cell(ri,2)
        lnk.hyperlink=_H(ref=lnk.coordinate,location=f"'{sh}'!A1")
        lnk.font=_F(name="Calibri",size=10,color="1A5276",underline="single")
        ws0.row_dimensions[ri].height=20
    ws0.row_dimensions[11].height=10
    _sec(ws0,12,2,4,"  SOURCE AND DISCLAIMER",SLATE,WHITE,11)
    for ri,(lbl,val) in enumerate([
        ("Source","UK Information Commissioner's Office - ico.org.uk"),
        ("Dataset","Data Security Incident Trends (2019 - Q4 2025)"),
        ("Generated",_dt.now().strftime("%d %B %Y, %H:%M")),
        ("Disclaimer","Engineered features are analytical constructs; not ICO classifications."),
        ("Reporting","ico.org.uk/for-organisations/report-a-breach/")],13):
        bg=LBG if ri%2==0 else WHITE
        c1=ws0.cell(ri,2,lbl);c1.font=_fnt(bold=True,size=10,color=NAVY)
        c1.fill=_fl(bg);c1.alignment=_lft();c1.border=_thin()
        ws0.merge_cells(start_row=ri,start_column=3,end_row=ri,end_column=4)
        c2=ws0.cell(ri,3,val);c2.font=_fnt(size=10);c2.fill=_fl(WHITE)
        c2.alignment=_lft();c2.border=_thin();ws0.row_dimensions[ri].height=20

    # Sheet 2: Data
    ws1=wb.create_sheet("Data");ws1.sheet_view.showGridLines=False
    avail=[c for c in ALL_C if c in df_export.columns]
    ws1.merge_cells(f"A1:{_gcl(len(avail))}1")
    c=ws1.cell(1,1,"ICO Breach Data - Filtered Export with Engineered Features")
    c.font=_F(name="Calibri",bold=True,size=13,color=WHITE);c.fill=_fl(NAVY)
    c.alignment=_A(horizontal="left",vertical="center",indent=2);ws1.row_dimensions[1].height=30
    raw_avail=[x for x in avail if x in RAW_C]
    eng_avail=[x for x in avail if x in ENG_C]
    if raw_avail:
        ws1.merge_cells(start_row=2,start_column=1,end_row=2,end_column=len(raw_avail))
        r=ws1.cell(2,1,"  ORIGINAL ICO FIELDS")
        r.font=_fnt(bold=True,size=10,color=WHITE);r.fill=_fl(SLATE)
        r.alignment=_A(horizontal="left",vertical="center",indent=1);ws1.row_dimensions[2].height=20
    if eng_avail:
        ws1.merge_cells(start_row=2,start_column=len(raw_avail)+1,end_row=2,end_column=len(avail))
        r2=ws1.cell(2,len(raw_avail)+1,"  ENGINEERED FEATURES - derived by this dashboard")
        r2.font=_fnt(bold=True,size=10,color=WHITE);r2.fill=_fl(PURPLE)
        r2.alignment=_A(horizontal="left",vertical="center",indent=1)
    for ci,col in enumerate(avail,1):
        c=ws1.cell(3,ci,CDN.get(col,col))
        bg=TEAL if col in RAW_C else PURPLE
        c.font=_fnt(bold=True,size=10,color=WHITE);c.fill=_fl(bg)
        c.alignment=_ctr();c.border=_thickb()
    ws1.row_dimensions[3].height=22
    num_c={'Is_Cyber','Is_High_Impact','Impact_Score','Is_Special_Category','Within_72hrs','Severity_Score','Year'}
    tier_c={"High":RED,"Medium":AMB,"Low":GREEN}
    cat_c={"Cyber":RED,"Non Cyber":CYAN}
    for ri,(idx,row) in enumerate(df_export[avail].iterrows(),4):
        rbg=LBG if ri%2==0 else WHITE
        for ci,col in enumerate(avail,1):
            val=row[col]
            c=ws1.cell(ri,ci,val)
            if col=="Incident_Category": c.fill=_fl(cat_c.get(str(val),rbg))
            elif col=="Sector_Risk_Tier": c.fill=_fl(tier_c.get(str(val),rbg))
            else: c.fill=_fl(rbg)
            c.font=_fnt(size=10)
            c.alignment=_rgt() if col in num_c else _lft()
            c.border=_thin()
        ws1.row_dimensions[ri].height=16
    ws1.freeze_panes="A4"
    ldr=3+len(df_export);lcl=_gcl(len(avail))
    try:
        t=_T(displayName="ICOData",ref=f"A3:{lcl}{ldr}")
        t.tableStyleInfo=_TSI(name="TableStyleMedium9",showRowStripes=True)
        ws1.add_table(t)
    except Exception:
        pass
    if "Severity_Score" in avail:
        sci=avail.index("Severity_Score")+1;sl=_gcl(sci)
        ws1.conditional_formatting.add(f"{sl}4:{sl}{ldr}",_CSR(
            start_type="min",start_color="63BE7B",
            mid_type="percentile",mid_value=50,mid_color="FFEB84",
            end_type="max",end_color="F8696B"))
    if "Impact_Score" in avail:
        ii=avail.index("Impact_Score")+1;il=_gcl(ii)
        ws1.conditional_formatting.add(f"{il}4:{il}{ldr}",_DBR(start_type="min",end_type="max",color="4472C4"))
    for ci in range(1,len(avail)+1): _aw(ws1,ci)

    # Sheet 3: Dictionary
    ws2=wb.create_sheet("Dictionary");ws2.sheet_view.showGridLines=False
    for col,w in [("A",3),("B",22),("C",26),("D",50),("E",14),("F",28),("G",46),("H",22)]:
        ws2.column_dimensions[col].width=w
    ws2.merge_cells("B2:H2")
    c=ws2.cell(2,2,"Data Dictionary - Field Definitions and UK GDPR Context")
    c.font=_fnt(bold=True,size=15,color=WHITE);c.fill=_fl(NAVY)
    c.alignment=_A(horizontal="left",vertical="center",indent=2);ws2.row_dimensions[2].height=38
    DH=["Internal Name","Display Name","Definition","Data Type","Example Values","Notes","Source"]
    _sec(ws2,4,2,8,"  ORIGINAL ICO FIELDS",TEAL,WHITE,11)
    for ci,h in enumerate(DH,2): _hdr(ws2.cell(5,ci,h),bg=TEAL); ws2.row_dimensions[5].height=22
    or2=[
        ("BI_Reference","Breach ID","Unique reference per breach report.","Text","BI1, BI100","Anonymised.","ICO dataset"),
        ("Year","Year","Calendar year of notification.","Integer","2019-2025","Notification year.","ICO dataset"),
        ("Quarter","Quarter","Qtr1=Jan-Mar, Qtr2=Apr-Jun, Qtr3=Jul-Sep, Qtr4=Oct-Dec.","Text","Qtr 1","Not incident quarter.","ICO dataset"),
        ("Sector","Sector","Reporting organisation sector per ICO classification.","Text","Health, Legal","Labelling varied historically.","ICO Glossary"),
        ("Data_Subject_Type","Data Subjects","Category of affected individuals (UK GDPR Art.4).","Text","Customers","One breach may have multiple rows.","UK GDPR Art.4(1)"),
        ("Data_Type","Data Category","Type of personal data. Art.9 = stricter obligations.","Text","Health data","Special category triggers Art.9.","UK GDPR Art.9"),
        ("Incident_Category","Breach Category","ICO typology: Cyber (malicious) or Non Cyber.","Text","Cyber","ICO internal classification.","ICO Glossary"),
        ("Incident_Type","Breach Type","Specific mechanism (ransomware, wrong recipient, etc.).","Text","Ransomware","Most granular breach description.","ICO Glossary"),
        ("Decision_Taken","ICO Decision","Regulatory response after harm risk assessment.","Text","No Further Action","Reflects ICO enforcement approach.","ICO Guidance"),
        ("No_Data_Subjects_Affected","People Affected","Estimated affected individuals as banded range.","Text (ordinal)","1 to 9","Estimate at notification.","UK GDPR Art.33"),
        ("Time_Taken_to_Report","Reporting Time","Hours from discovery to ICO notification.","Text","0 to 24 hours","Art.33 requires max 72hrs.","UK GDPR Art.33"),
    ]
    for ri,row in enumerate(or2,6):
        bg=LBG if ri%2==0 else WHITE
        for ci,val in enumerate(row,2):
            c=ws2.cell(ri,ci,val);c.fill=_fl(bg);c.font=_fnt(size=10)
            c.alignment=_A(horizontal="left",vertical="center",wrap_text=True,indent=1);c.border=_thin()
        ws2.row_dimensions[ri].height=40
    es=6+len(or2)+2;_sec(ws2,es,2,8,"  ENGINEERED FEATURES",PURPLE,WHITE,11)
    for ci,h in enumerate(DH,2): _hdr(ws2.cell(es+1,ci,h),bg=PURPLE); ws2.row_dimensions[es+1].height=22
    en2=[
        ("Is_Cyber","Is Cyber? (0/1)","1 if Incident_Category == Cyber.","Integer (0/1)","0, 1","Enables numeric ops.","Derived"),
        ("Is_High_Impact","High Impact? (0/1)","1 if People Affected >= 1,000.","Integer (0/1)","0, 1","Art.33(3) high-risk threshold.","Derived / UK GDPR Art.33"),
        ("Impact_Score","Impact Score (1-6)","Ordinal: 1-to-9=1 through Over-100k=6.","Integer (1-6)","1, 3, 6","Numeric proxy for scale.","Derived"),
        ("Is_Special_Category","Special Category? (0/1)","1 if Data_Type contains Art.9 keyword.","Integer (0/1)","0, 1","Health/biometric/racial/political data.","Derived / UK GDPR Art.9"),
        ("Within_72hrs","Within 72hrs? (0/1)","1 if reported within 72hrs.","Integer (0/1)","0, 1","Art.33 compliance signal.","Derived / UK GDPR Art.33"),
        ("Severity_Score","Severity Score (0-11)","Is_Cyber*3 + Impact_Score + Is_Special_Category*2.","Integer (0-11)","0, 5, 11","Composite risk. Not an ICO metric.","Derived"),
        ("Sector_Risk_Tier","Sector Risk Tier","Cyber rates ranked into High/Medium/Low tertiles.","Text","High, Medium, Low","Relative. Not an ICO designation.","Derived"),
    ]
    for ri,row in enumerate(en2,es+2):
        bg=PUR if ri%2==0 else WHITE
        for ci,val in enumerate(row,2):
            c=ws2.cell(ri,ci,val);c.fill=_fl(bg);c.font=_fnt(size=10)
            c.alignment=_A(horizontal="left",vertical="center",wrap_text=True,indent=1);c.border=_thin()
        ws2.row_dimensions[ri].height=42
    ws2.freeze_panes="B6"

    # Sheet 4: Methods
    ws3=wb.create_sheet("Methods");ws3.sheet_view.showGridLines=False
    for col,w in [("A",3),("B",24),("C",50),("D",38),("E",28)]: ws3.column_dimensions[col].width=w
    ws3.merge_cells("B2:E2")
    c=ws3.cell(2,2,"Data Cleaning and Feature Engineering - Methods Reference")
    c.font=_fnt(bold=True,size=15,color=WHITE);c.fill=_fl(NAVY)
    c.alignment=_A(horizontal="left",vertical="center",indent=2);ws3.row_dimensions[2].height=38
    MH=["Step / Feature","What was done","Python technique","Why it matters"]
    _sec(ws3,4,2,5,"  DATA CLEANING STEPS",TEAL,WHITE,11)
    for ci,h in enumerate(MH,2): _hdr(ws3.cell(5,ci,h),bg=TEAL); ws3.row_dimensions[5].height=22
    cl=[
        ("1. Column renaming","Column names stripped and renamed to snake_case.",
         "df.columns=[c.strip() for c in df.columns]; df.rename(columns={...})","Prevents key errors. Standard for production pipelines."),
        ("2. Whitespace trimming","Leading/trailing spaces removed from text fields.",
         "df[col].str.strip() at load time","ICO exports sometimes contain trailing spaces."),
        ("3. Ordinal encoding","People Affected converted to ordered Categorical.",
         "pd.Categorical(col,categories=bands,ordered=True)","Without ordering, charts sort alphabetically."),
        ("4. Date construction","Date column built from Year + Quarter midpoint month.",
         "q_map={'Qtr 1':2,'Qtr 2':5,'Qtr 3':8,'Qtr 4':11}; pd.to_datetime(yr+'-'+mo+'-01')","Raw data has no date field."),
        ("5. Missing value handling","No imputation on original fields. Blanks preserved.",
         "df.isnull().sum() displayed in dashboard. No fillna() on originals.","Imputing without domain knowledge introduces bias."),
        ("6. 2025 annotation","Full Q1-Q4 2025 data included. Annotated in trend charts.",
         "Expander notes in Trends and Data Quality tabs.","Prevents misreading of year-on-year comparisons."),
        ("7. No deduplication","One breach = multiple rows per data subject type. Preserved.",
         "No df.drop_duplicates(). Intentional design.","Deduplicating would lose data type granularity."),
    ]
    for ri,row in enumerate(cl,6):
        bg=LBG if ri%2==0 else WHITE
        for ci,val in enumerate(row,2):
            c=ws3.cell(ri,ci,val);c.fill=_fl(bg);c.font=_fnt(size=10)
            c.alignment=_A(horizontal="left",vertical="center",wrap_text=True,indent=1);c.border=_thin()
        ws3.row_dimensions[ri].height=54
    fs=6+len(cl)+2;_sec(ws3,fs,2,5,"  FEATURE ENGINEERING METHODS",PURPLE,WHITE,11)
    for ci,h in enumerate(["Feature","Construction method","Python code","Analytical purpose"],2):
        _hdr(ws3.cell(fs+1,ci,h),bg=PURPLE); ws3.row_dimensions[fs+1].height=22
    fe=[
        ("Is_Cyber","Binary flag from ICO Incident_Category.","(df['Incident_Category']=='Cyber').astype(int)","Enables KPI calculation and model target."),
        ("Is_High_Impact","Binary: 1 if People Affected >= 1,000.","df['No_Data_Subjects_Affected'].isin(['1k to 10k','10k to 100k','Over 100k']).astype(int)","Flags Art.33(3) high-risk breaches."),
        ("Impact_Score","Ordinal 1-6 from band strings.","band_score={'1 to 9':1,...,'Over 100k':6}; df['No_Data_Subjects_Affected'].map(band_score)","Numeric proxy for impact scale."),
        ("Is_Special_Category","Keyword scan for UK GDPR Art.9 categories.","keywords=['health','racial','ethnic','biometric','genetic','sexual','religion','political','criminal']; df['Data_Type'].str.lower().apply(lambda x:int(any(k in x for k in keywords)))","Flags highest regulatory risk data."),
        ("Within_72hrs","Parse time field for compliance signals.","df['Time_Taken_to_Report'].str.lower().apply(lambda x:1 if any(t in x for t in ['0 to 24','24 to 48','48 to 72']) else 0)","UK GDPR Art.33 compliance signal."),
        ("Severity_Score","Composite 3-dimension risk metric.","df['Is_Cyber']*3 + df['Impact_Score'] + df['Is_Special_Category']*2","Range 0-11. Higher = more potentially harmful."),
        ("Sector_Risk_Tier","Sector cyber rates ranked into tertiles.","rate=df.groupby('Sector')['Is_Cyber'].mean(); q33,q66=rate.quantile(0.33),rate.quantile(0.66)","Relative risk tier. Not ICO-designated."),
    ]
    for ri,row in enumerate(fe,fs+2):
        bg=PUR if ri%2==0 else WHITE
        for ci,val in enumerate(row,2):
            c=ws3.cell(ri,ci,val);c.fill=_fl(bg);c.font=_fnt(size=10)
            c.alignment=_A(horizontal="left",vertical="center",wrap_text=True,indent=1);c.border=_thin()
        ws3.row_dimensions[ri].height=62
    ws3.freeze_panes="B6"

    # Sheet 5: Stats
    ws4=wb.create_sheet("Stats");ws4.sheet_view.showGridLines=False
    for col,w in [("A",3),("B",34),("C",20),("D",16),("E",36),("F",18)]: ws4.column_dimensions[col].width=w
    ws4.merge_cells("B2:F2")
    c=ws4.cell(2,2,"Summary Statistics - ICO Breach Data (Filtered Selection)")
    c.font=_fnt(bold=True,size=15,color=WHITE);c.fill=_fl(NAVY)
    c.alignment=_A(horizontal="left",vertical="center",indent=2);ws4.row_dimensions[2].height=38
    _sec(ws4,4,2,6,"  KEY PERFORMANCE INDICATORS",TEAL,WHITE,11)
    for ci,h in enumerate(["Metric","Value","Format","Interpretation","Source"],2):
        _hdr(ws4.cell(5,ci,h),bg=TEAL); ws4.row_dimensions[5].height=22
    ic=df_export["Is_Cyber"] if "Is_Cyber" in df_export.columns else pd.Series([0]*len(df_export))
    hi=df_export["Is_High_Impact"] if "Is_High_Impact" in df_export.columns else pd.Series([0]*len(df_export))
    sv=df_export["Severity_Score"] if "Severity_Score" in df_export.columns else pd.Series([0]*len(df_export))
    im=df_export["Impact_Score"] if "Impact_Score" in df_export.columns else pd.Series([0]*len(df_export))
    sc_f=df_export["Is_Special_Category"] if "Is_Special_Category" in df_export.columns else pd.Series([0]*len(df_export))
    w7=df_export["Within_72hrs"] if "Within_72hrs" in df_export.columns else pd.Series([0]*len(df_export))
    st_c=df_export["Sector"] if "Sector" in df_export.columns else pd.Series(["Unknown"]*len(df_export))
    yr_c=df_export["Year"] if "Year" in df_export.columns else pd.Series([0]*len(df_export))
    srt=df_export["Sector_Risk_Tier"] if "Sector_Risk_Tier" in df_export.columns else pd.Series(["Low"]*len(df_export))
    kps=[
        ("Total Breach Reports",len(df_export),"#,##0","All rows in this filtered export","ICO dataset"),
        ("Cyber Breaches",int(ic.sum()),"#,##0","ICO-classified as cyber origin","Derived: Is_Cyber"),
        ("Cyber Breach Rate",round(float(ic.mean()),4),"0.0%","34% is the full-dataset average","Derived: Is_Cyber"),
        ("High-Impact Breaches",int(hi.sum()),"#,##0","Affected 1,000+ individuals","Derived: Is_High_Impact"),
        ("Avg. Severity Score",round(float(sv.mean()),2),"0.00","Scale 0-11","Derived: Severity_Score"),
        ("Special Category Breaches",int(sc_f.sum()),"#,##0","UK GDPR Art.9 data involved","Derived: Is_Special_Category"),
        ("Within 72hrs (rate)",round(float(w7.mean()),4),"0.0%","Art.33 compliance signal","Derived: Within_72hrs"),
        ("Sectors represented",int(st_c.nunique()),"0","Unique sectors in selection","ICO dataset"),
        ("High-Risk Sectors",int((srt=="High").sum()),"0","Top tertile by cyber rate","Derived: Sector_Risk_Tier"),
        ("Years covered",str(int(yr_c.min()))+" to "+str(int(yr_c.max())),"@","Date range of filtered data","ICO dataset"),
    ]
    for ri,(lbl,val,fmt,interp,src) in enumerate(kps,6):
        bg=LBG if ri%2==0 else WHITE
        c1=ws4.cell(ri,2,lbl);c1.font=_fnt(bold=True,size=10);c1.fill=_fl(bg);c1.alignment=_lft();c1.border=_thin()
        c2=ws4.cell(ri,3,val);c2.number_format=fmt
        c2.font=_fnt(bold=True,size=11,color=NAVY);c2.fill=_fl(CYAN if ri%2==0 else "EBF5FB")
        c2.alignment=_ctr();c2.border=_thin()
        c3=ws4.cell(ri,4,fmt.replace("@","Text").replace("#,##0","Integer").replace("0.0%","Pct").replace("0.00","Decimal"))
        c3.font=_fnt(size=10);c3.fill=_fl(bg);c3.alignment=_ctr();c3.border=_thin()
        c4=ws4.cell(ri,5,interp);c4.font=_fnt(size=10);c4.fill=_fl(bg)
        c4.alignment=_A(horizontal="left",vertical="center",wrap_text=True,indent=1);c4.border=_thin()
        c5=ws4.cell(ri,6,src);c5.font=_fnt(size=10,italic=True);c5.fill=_fl(bg);c5.alignment=_ctr();c5.border=_thin()
        ws4.row_dimensions[ri].height=22
    if "Year" in df_export.columns and "Is_Cyber" in df_export.columns:
        ys=6+len(kps)+3;_sec(ws4,ys,2,6,"  BREACH REPORTS BY YEAR",AMBER,WHITE,11)
        for ci,h in enumerate(["Year","Total","Cyber","Non-Cyber","Cyber Rate"],2):
            _hdr(ws4.cell(ys+1,ci,h),bg=AMBER); ws4.row_dimensions[ys+1].height=22
        yg=df_export.groupby("Year").agg(T=("Is_Cyber","count"),C=("Is_Cyber","sum"),N=("Is_Cyber",lambda x:(x==0).sum())).reset_index()
        yg["R"]=yg["C"]/yg["T"]
        for ri,(idx_r,row) in enumerate(yg.iterrows(),ys+2):
            bg=AMB if ri%2==0 else WHITE
            for ci,(val,fmt) in enumerate([(int(row["Year"]),"0"),(int(row["T"]),"#,##0"),(int(row["C"]),"#,##0"),(int(row["N"]),"#,##0"),(row["R"],"0.0%")],2):
                c=ws4.cell(ri,ci,val);c.number_format=fmt;c.fill=_fl(bg)
                c.font=_fnt(size=10);c.alignment=_ctr();c.border=_thin()
            ws4.row_dimensions[ri].height=18
    ws4.freeze_panes="B6"

    buf=_IO();wb.save(buf);buf.seek(0);return buf.getvalue()

# ── LOAD + FEATURE ENGINEER ────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def load_data():
    df = pd.read_csv("ico_raw.csv")
    df.columns = [c.strip() for c in df.columns]
    df = df.rename(columns={
        "BI Reference":"BI_Reference","Year":"Year","Quarter":"Quarter",
        "Data Subject Type":"Data_Subject_Type","Data Type":"Data_Type",
        "Decision Taken":"Decision_Taken","Incident Category":"Incident_Category",
        "Incident Type":"Incident_Type","No. Data Subjects Affected":"No_Data_Subjects_Affected",
        "Sector":"Sector","Time Taken to Report":"Time_Taken_to_Report",
    })
    bands = ["1 to 9","10 to 99","100 to 1k","1k to 10k","10k to 100k","Over 100k"]
    df["No_Data_Subjects_Affected"] = pd.Categorical(df["No_Data_Subjects_Affected"],categories=bands,ordered=True)
    q_map = {"Qtr 1":2,"Qtr 2":5,"Qtr 3":8,"Qtr 4":11}
    df["Month"]   = df["Quarter"].map(q_map).fillna(2).astype(int)
    df["Date"]    = pd.to_datetime(df["Year"].astype(str)+"-"+df["Month"].astype(str).str.zfill(2)+"-01")
    df["YearQtr"] = df["Year"].astype(str)+" "+df["Quarter"].astype(str)
    band_score    = {"1 to 9":1,"10 to 99":2,"100 to 1k":3,"1k to 10k":4,"10k to 100k":5,"Over 100k":6}
    sc_kw         = ["health","racial","ethnic","biometric","genetic","sexual","religion","political","criminal"]
    df["Is_Cyber"]            = (df["Incident_Category"]=="Cyber").astype(int)
    df["Is_High_Impact"]      = df["No_Data_Subjects_Affected"].astype(str).isin(["1k to 10k","10k to 100k","Over 100k"]).astype(int)
    df["Impact_Score"]        = df["No_Data_Subjects_Affected"].astype(str).map(band_score).fillna(0).astype(int)
    df["Is_Special_Category"] = df["Data_Type"].fillna("").str.lower().apply(lambda x:int(any(k in x for k in sc_kw)))
    if "Time_Taken_to_Report" in df.columns:
        df["Within_72hrs"]    = df["Time_Taken_to_Report"].fillna("").str.lower().apply(
            lambda x:1 if any(t in x for t in ["0 to 24","24 to 48","48 to 72","within 72","<72"]) else 0)
    else:
        df["Within_72hrs"]    = 0
    df["Severity_Score"]      = df["Is_Cyber"]*3+df["Impact_Score"]+df["Is_Special_Category"]*2
    return df

df_full = load_data()

def compute_sector_tiers(df):
    rate = df.groupby("Sector")["Is_Cyber"].mean()
    q33,q66 = rate.quantile(0.33),rate.quantile(0.66)
    return df["Sector"].map(rate).apply(lambda r:"High" if r>=q66 else("Medium" if r>=q33 else "Low"))

df_full["Sector_Risk_Tier"] = compute_sector_tiers(df_full)

@st.cache_resource(show_spinner=False)
def get_model_results():
    return train_models(df_full)

model_results = get_model_results()
best_name = max(model_results, key=lambda k: model_results[k]["auc"])

@st.cache_resource(show_spinner=False)
def train_models(df):
    data = df[df["Incident_Category"].isin(["Cyber","Non Cyber"])].copy()
    data["No_Data_Subjects_Affected"] = data["No_Data_Subjects_Affected"].astype(str)
    y     = data["Is_Cyber"]
    feats = ["Sector","Data_Subject_Type","Data_Type","Incident_Type",
             "No_Data_Subjects_Affected","Time_Taken_to_Report","Year",
             "Is_Special_Category","Impact_Score","Within_72hrs"]
    cats  = ["Sector","Data_Subject_Type","Data_Type","Incident_Type",
             "No_Data_Subjects_Affected","Time_Taken_to_Report"]
    nums  = ["Year","Is_Special_Category","Impact_Score","Within_72hrs"]
    X     = data[feats].copy()
    Xtr,Xte,ytr,yte = train_test_split(X,y,stratify=y,test_size=0.2,random_state=42)
    try:
        ohe = OneHotEncoder(handle_unknown="ignore",sparse_output=False)
    except TypeError:
        ohe = OneHotEncoder(handle_unknown="ignore",sparse=False)
    preproc = ColumnTransformer([
        ("cat",Pipeline([("imp",SimpleImputer(strategy="most_frequent")),("ohe",ohe)]),cats),
        ("num",SimpleImputer(strategy="median"),nums),
    ])
    clfs = [
        ("Logistic Regression",LogisticRegression(max_iter=1000,random_state=42)),
        ("Random Forest",RandomForestClassifier(n_estimators=150,max_depth=8,random_state=42,n_jobs=-1)),
        ("Gradient Boosting",GradientBoostingClassifier(n_estimators=100,max_depth=4,random_state=42)),
    ]
    results = {}
    for name,clf in clfs:
        pipe  = Pipeline([("preproc",preproc),("clf",clf)])
        pipe.fit(Xtr,ytr)
        yp    = pipe.predict(Xte)
        yprob = pipe.predict_proba(Xte)[:,1]
        rep   = classification_report(yte,yp,output_dict=True)
        auc   = roc_auc_score(yte,yprob)
        fpr,tpr,_ = roc_curve(yte,yprob)
        cm    = confusion_matrix(yte,yp)
        entry = {"pipe":pipe,"report":rep,"auc":auc,"fpr":fpr,"tpr":tpr,"cm":cm}
        if name=="Random Forest":
            try:
                ohe_s = pipe.named_steps["preproc"].named_transformers_["cat"].named_steps["ohe"]
                fn    = list(ohe_s.get_feature_names_out(cats))+nums
                fi    = pd.DataFrame({"Feature":fn,"Importance":clf.feature_importances_}).sort_values("Importance",ascending=False).head(20)
                entry["fi"] = fi
            except Exception:
                entry["fi"] = None
        results[name] = entry
    return results

def insights(df):
    if df.empty:
        return ["No data matches the current filters."]
    out=[]
    cyber=df["Is_Cyber"]==1
    if cyber.any():
        ts=df[cyber]["Sector"].value_counts().idxmax()
        tc=df[cyber]["Sector"].value_counts().iloc[0]
        out.append("🔴 **"+ts+"** recorded the most cyber breaches ("+f"{tc:,}"+"). The ICO defines cyber breaches as those involving malicious third-party actors such as ransomware or phishing.")
    pct=round(cyber.mean()*100,1)
    base=round(df_full["Is_Cyber"].mean()*100,1)
    flag="**above** the full-dataset average of "+str(base)+"%" if pct>base else "**within** the typical range"
    out.append("📊 Cyber breaches represent **"+str(pct)+"%** of reports — "+flag+".")
    hi=round(df["Is_High_Impact"].mean()*100,1)
    if hi>10:
        out.append("⚠️ **"+str(hi)+"%** of breaches affected 1,000+ people — above the 10% threshold. The ICO requires organisations to assess and document the risk of harm to affected individuals.")
    sc_rate=round(df["Is_Special_Category"].mean()*100,1)
    if sc_rate>0:
        out.append("🏥 **"+str(sc_rate)+"%** involved **special category data** (health, biometric, racial/ethnic, etc.). Under UK GDPR Article 9, this carries the strictest protection obligations.")
    avg_sev=round(df["Severity_Score"].mean(),2)
    out.append("📈 Average **Severity Score**: **"+str(avg_sev)+" / 11** (cyber=3pts, impact band=1-6pts, special category=2pts).")
    tt=df["Incident_Type"].value_counts().idxmax()
    ttc=df["Incident_Type"].value_counts().iloc[0]
    out.append("📌 Most frequent breach type: **"+tt+"** ("+f"{ttc:,}"+" reports).")
    hi_sec=(df.groupby("Sector")["Sector_Risk_Tier"].first()=="High").sum()
    if hi_sec>0:
        top3=", ".join(df.groupby("Sector")["Sector_Risk_Tier"].first()[df.groupby("Sector")["Sector_Risk_Tier"].first()=="High"].index.tolist()[:3])
        out.append("🎯 **"+str(hi_sec)+" sectors** are in the High Cyber Risk Tier — including "+top3+".")
    return out

# ── SIDEBAR ────────────────────────────────────────────────────────────────────
st.sidebar.markdown(
    '<div style="text-align:center;padding:12px 0 8px;">'
    '<img src="https://ico.org.uk/media/about-the-ico/images/ico-logo-2019.png" '
    'width="160" style="max-width:100%;border-radius:4px;" onerror="this.style.display=\'none\'" />'
    '<p style="color:#94a3b8;font-size:0.75rem;margin-top:6px;">Information Commissioner\'s Office</p>'
    '</div>', unsafe_allow_html=True)
st.sidebar.markdown("## Filters")
st.sidebar.caption("Filters apply to all tabs.")
years      = sorted(df_full["Year"].dropna().unique())
sectors    = sorted(df_full["Sector"].dropna().unique())
categories = sorted(df_full["Incident_Category"].dropna().unique())
year_sel   = st.sidebar.multiselect("Year",years,default=years)
sector_sel = st.sidebar.multiselect("Sector",sectors,default=sectors)
cat_sel    = st.sidebar.multiselect("Breach category",categories,default=categories)
filtered   = df_full[df_full["Year"].isin(year_sel)&df_full["Sector"].isin(sector_sel)&df_full["Incident_Category"].isin(cat_sel)].copy()
st.sidebar.markdown("---")
st.sidebar.markdown(
    '<div style="background:#1e293b;border-radius:8px;padding:12px;">'
    '<div style="font-size:0.7rem;color:#94a3b8;text-transform:uppercase;letter-spacing:.06em;">Records in view</div>'
    '<div style="font-size:1.4rem;font-weight:700;color:#f1f5f9;">'+f"{len(filtered):,}"+'</div>'
    '</div>', unsafe_allow_html=True)
st.sidebar.markdown("---")
st.sidebar.markdown("**Source:** [ICO Breach Trends](https://ico.org.uk/action-weve-taken/complaints-and-concerns-data-sets/data-security-incident-trends/)  \nData: 2019 – Q4 2025")

# ── HEADER ─────────────────────────────────────────────────────────────────────
st.markdown(
    '<div class="hero"><h1>🔐 ICO Data Security Incident Explorer</h1>'
    '<p>Interactive analysis of personal data breaches self-reported to the UK '
    '<strong>Information Commissioner\'s Office (ICO)</strong> — 2019 to Q4 2025. '
    'The ICO publishes this data to help organisations understand breach trends and meet their UK GDPR obligations. '
    'Use the sidebar filters to focus on a specific time period, sector, or breach category.</p></div>',
    unsafe_allow_html=True)

tabs = st.tabs(["📊 Overview","📈 Trends","🏢 Sector Analysis","⚠️ Impact & Severity",
                "🧪 Feature Insights","🔍 Data Quality","🤖 Predictive Model",
                "🔮 Risk Predictor","💡 Key Insights","📋 About the Data"])

# --- TAB 0 — OVERVIEW ----------------------------------------------------------
with tabs[0]:
    if filtered.empty:
        st.warning("No data matches. Adjust sidebar filters.")
    else:
        total     = len(filtered)
        pct_cyber = round(filtered["Is_Cyber"].mean()*100,1)
        pct_hi    = round(filtered["Is_High_Impact"].mean()*100,1)
        yr_range  = (str(min(year_sel))+" – "+str(max(year_sel))) if year_sel else "N/A"
        top_s     = filtered["Sector"].value_counts().idxmax()
        n_secs    = filtered["Sector"].nunique()
        avg_sev   = round(filtered["Severity_Score"].mean(),1)
        sc_pct    = round(filtered["Is_Special_Category"].mean()*100,1)
        hi_risk   = (filtered.groupby("Sector")["Sector_Risk_Tier"].first()=="High").sum()

        st.markdown('<div class="section-tag">Headline Figures</div>',unsafe_allow_html=True)
        r1c1,r1c2,r1c3,r1c4 = st.columns(4)
        r1c1.markdown(kpi("Total Breach Reports",f"{total:,}","All self-reported to ICO"),unsafe_allow_html=True)
        r1c2.markdown(kpi("Cyber Breaches",str(pct_cyber)+"%","Malicious/technical origin"),unsafe_allow_html=True)
        r1c3.markdown(kpi("High-Impact Breaches",str(pct_hi)+"%","1,000+ people affected"),unsafe_allow_html=True)
        r1c4.markdown(kpi("Period Covered",yr_range,"Financial years"),unsafe_allow_html=True)
        st.markdown("<div style='height:12px'></div>",unsafe_allow_html=True)
        r2c1,r2c2,r2c3,r2c4 = st.columns(4)
        r2c1.markdown(kpi("Avg. Severity Score",str(avg_sev)+" / 11","Cyber+impact+special data"),unsafe_allow_html=True)
        r2c2.markdown(kpi("Special Category Data",str(sc_pct)+"%","Art.9 UK GDPR — heightened risk"),unsafe_allow_html=True)
        r2c3.markdown(kpi("High-Risk Sectors",str(hi_risk),"Elevated cyber breach rate"),unsafe_allow_html=True)
        r2c4.markdown(kpi("Sectors in Selection",str(n_secs),"Most reported: "+top_s[:18]),unsafe_allow_html=True)
        note("What do these figures mean?",
             "- **Total Breach Reports** — all personal data breach notifications after filters.\n"
             "- **Cyber Breaches** — share classified by ICO as cyber (malicious third-party origin).\n"
             "- **High-Impact** — breaches affecting 1,000+ people.\n"
             "- **Avg. Severity Score** — 0–11: 3pts for cyber, 1–6pts for impact band, 2pts for special category data.\n"
             "- **Special Category Data** — health, biometric, racial/ethnic, religious, criminal (UK GDPR Article 9).\n"
             "- **High-Risk Sectors** — top third by cyber breach rate across all sectors.")
        st.markdown("---")
        st.markdown('<div class="section-tag">What the data tells us</div>',unsafe_allow_html=True)
        for ins in insights(filtered):
            st.info(ins)
        st.markdown("---")
        st.markdown('<div class="section-tag">Breach Trends at a Glance</div>',unsafe_allow_html=True)
        ch1,ch2 = st.columns([3,2])
        with ch1:
            st.markdown("#### Cyber vs Non-Cyber breach reports over time")
            tdf=(filtered.groupby(["Date","Incident_Category"]).size().reset_index(name="Reports").sort_values("Date"))
            fig_t=px.line(tdf,x="Date",y="Reports",color="Incident_Category",markers=True,
                          color_discrete_map={"Cyber":C_CYBER,"Non Cyber":C_NONCYBER},template=TPL,height=320)
            fig_t.update_layout(legend_title="Breach Category",margin=dict(t=10,b=10))
            st.plotly_chart(fig_t,use_container_width=True)
            note("How to read this chart","Red = cyber (malicious). Blue = non-cyber (human error, physical loss). A rising red line indicates growing cyber threat activity or improved reporting.")
        with ch2:
            st.markdown("#### Breach category breakdown")
            cat_bar=filtered["Incident_Category"].value_counts().reset_index()
            cat_bar.columns=["Category","Count"]
            cat_bar["Share (%)"]=round(cat_bar["Count"]/cat_bar["Count"].sum()*100,1)
            fig_cb=px.bar(cat_bar,x="Category",y="Count",color="Category",text="Share (%)",
                          color_discrete_map={"Cyber":C_CYBER,"Non Cyber":C_NONCYBER},template=TPL,height=320)
            fig_cb.update_traces(texttemplate="%{text}%",textposition="outside")
            fig_cb.update_layout(showlegend=False,margin=dict(t=10,b=10),yaxis_title="Breach Reports")
            st.plotly_chart(fig_cb,use_container_width=True)
            note("How to read this chart","Bar height = absolute volume. Percentage label = proportional split. Allows direct comparison of both volume and proportion.")
        st.markdown("---")
        st.markdown('<div class="section-tag">Year-on-Year Cyber Trend</div>',unsafe_allow_html=True)
        st.markdown("#### Annual cyber breach rate")
        yoy=filtered.groupby("Year").agg(Total=("Is_Cyber","count"),Cyber=("Is_Cyber","sum")).reset_index()
        yoy["Cyber Rate (%)"]=round(yoy["Cyber"]/yoy["Total"]*100,1)
        fig_yoy=make_subplots(specs=[[{"secondary_y":True}]])
        fig_yoy.add_trace(go.Bar(x=yoy["Year"],y=yoy["Cyber"],name="Cyber Breach Count",marker_color=C_CYBER,opacity=0.75),secondary_y=False)
        fig_yoy.add_trace(go.Scatter(x=yoy["Year"],y=yoy["Cyber Rate (%)"],name="Cyber Rate (%)",mode="lines+markers",line=dict(color=C_WARN,width=2.5),marker=dict(size=8)),secondary_y=True)
        fig_yoy.update_layout(template=TPL,height=360,legend=dict(orientation="h",y=-0.15),yaxis_title="Cyber Breaches",yaxis2_title="Cyber Breach Rate (%)",hovermode="x unified")
        st.plotly_chart(fig_yoy,use_container_width=True)
        note("How to read this dual-axis chart","Bars (left axis) = absolute cyber breach count. Line (right axis) = cyber rate as % of all breaches that year. If both rise together, cyber is growing in both volume and share.")

# --- TAB 1 — TRENDS ------------------------------------------------------------
with tabs[1]:
    st.markdown("### Breach Report Trends")
    st.info("This version uses full Q1–Q4 2025 data. All years are directly comparable.")
    if not filtered.empty:
        gran=st.selectbox("Group by",["Year","Quarter (Year + Qtr)"],key="gran")
        work=filtered.copy()
        work["TimeBucket"]=work["Year"].astype(str)+" "+work["Quarter"].astype(str) if gran=="Quarter (Year + Qtr)" else work["Year"].astype(str)
        trend=work.groupby(["TimeBucket","Incident_Category"]).size().reset_index(name="Reports")
        fig_tr=px.bar(trend,x="TimeBucket",y="Reports",color="Incident_Category",barmode="group",
                      color_discrete_map={"Cyber":C_CYBER,"Non Cyber":C_NONCYBER},template=TPL,height=400)
        fig_tr.update_layout(xaxis_title="Period",legend_title="Breach Category")
        st.plotly_chart(fig_tr,use_container_width=True)
        note("How to read this chart","Each pair of bars = one time period. Red = cyber, Blue = non-cyber. Comparing across periods shows whether cyber reporting is growing.")
        st.markdown("---")
        st.markdown("#### Cyber share as proportion of all breaches")
        share=work.groupby(["TimeBucket","Incident_Category"]).size().reset_index(name="Count")
        fig_a=px.area(share,x="TimeBucket",y="Count",color="Incident_Category",groupnorm="fraction",
                      color_discrete_map={"Cyber":C_CYBER,"Non Cyber":C_NONCYBER},template=TPL,height=340)
        fig_a.update_layout(yaxis_title="Proportion",yaxis_tickformat=".0%",legend_title="Breach Category")
        st.plotly_chart(fig_a,use_container_width=True)
        note("How to read this chart","Growing red area = cyber breaches rising as a share of all reported breaches.")
        st.markdown("---")
        st.markdown("#### Severity Score trend over time")
        sev_t=filtered.groupby("Year")["Severity_Score"].mean().round(2).reset_index()
        sev_t.columns=["Year","Avg Severity Score"]
        fig_sv=px.line(sev_t,x="Year",y="Avg Severity Score",markers=True,color_discrete_sequence=[C_PURPLE],template=TPL,height=320)
        st.plotly_chart(fig_sv,use_container_width=True)
        note("How to read this chart","Rising trend = breaches becoming more severe on average — more cyber-classified, affecting more people, or involving more sensitive data.")
        st.markdown("---")
        n_t=st.slider("Top breach types to show",3,8,5,key="n_types")
        top_types=filtered["Incident_Type"].value_counts().head(n_t).index.tolist()
        type_df=(filtered[filtered["Incident_Type"].isin(top_types)].groupby(["Date","Incident_Type"]).size().reset_index(name="Reports").sort_values("Date"))
        if not type_df.empty:
            fig_tt=px.line(type_df,x="Date",y="Reports",color="Incident_Type",markers=True,template=TPL,height=380)
            st.plotly_chart(fig_tt,use_container_width=True)
            note("How to read this chart","Each line = one breach type. Rising lines indicate that type is being reported more frequently over time.")

# --- TAB 2 — SECTOR ANALYSIS ---------------------------------------------------
with tabs[2]:
    st.markdown("### Breach Reports by Sector")
    if not filtered.empty:
        top_n=st.slider("Sectors to display",5,20,10,key="topn_s")
        col_a,col_b=st.columns(2)
        with col_a:
            sc=filtered.groupby(["Sector","Incident_Category"]).size().reset_index(name="Reports")
            tops=sc.groupby("Sector")["Reports"].sum().nlargest(top_n).index
            fig_s=px.bar(sc[sc["Sector"].isin(tops)],y="Sector",x="Reports",color="Incident_Category",barmode="stack",orientation="h",
                         color_discrete_map={"Cyber":C_CYBER,"Non Cyber":C_NONCYBER},template=TPL,height=520)
            fig_s.update_layout(yaxis=dict(categoryorder="total ascending"),legend_title="Breach Category")
            st.plotly_chart(fig_s,use_container_width=True)
            note("How to read this chart","Bar length = total reports. Red = cyber. Long red = predominantly cyber-driven sector.")
        with col_b:
            st.markdown("#### Cyber breach rate by sector")
            cr=filtered.groupby("Sector")["Is_Cyber"].mean().mul(100).round(1).reset_index()
            cr.columns=["Sector","Cyber Breach Rate (%)"]
            cr=cr.sort_values("Cyber Breach Rate (%)",ascending=False).head(20)
            tier_map=filtered.groupby("Sector")["Sector_Risk_Tier"].first().to_dict()
            cr["Risk Tier"]=cr["Sector"].map(tier_map).fillna("Medium")
            tier_colours={"High":C_CYBER,"Medium":C_WARN,"Low":C_NEUTRAL}
            fig_cr=px.bar(cr,x="Cyber Breach Rate (%)",y="Sector",orientation="h",color="Risk Tier",
                          color_discrete_map=tier_colours,template=TPL,height=520)
            fig_cr.update_layout(yaxis=dict(categoryorder="total ascending"),xaxis_title="% of sector reports classified as cyber")
            st.plotly_chart(fig_cr,use_container_width=True)
            note("How to read this chart","% of each sector's reports classified as cyber. Red = High risk tier (top third), Amber = Medium, Green = Low.")
        st.markdown("---")
        s_pick=st.selectbox("Drill into a sector",["All sectors"]+sorted(filtered["Sector"].dropna().unique().tolist()),key="sector_dd")
        sec_df=filtered if s_pick=="All sectors" else filtered[filtered["Sector"]==s_pick]
        d1,d2=st.columns(2)
        with d1:
            it=sec_df["Incident_Type"].value_counts().head(10).reset_index(); it.columns=["Breach Type","Reports"]
            fig_it=px.bar(it,x="Reports",y="Breach Type",orientation="h",template=TPL,height=360,title="Most common breach types",color_discrete_sequence=[C_CYBER])
            fig_it.update_layout(yaxis=dict(categoryorder="total ascending")); st.plotly_chart(fig_it,use_container_width=True)
        with d2:
            dt=sec_df["Data_Type"].value_counts().head(10).reset_index(); dt.columns=["Data Category","Reports"]
            fig_dt=px.bar(dt,x="Reports",y="Data Category",orientation="h",template=TPL,height=360,title="Data categories affected",color_discrete_sequence=[C_NEUTRAL])
            fig_dt.update_layout(yaxis=dict(categoryorder="total ascending")); st.plotly_chart(fig_dt,use_container_width=True)
            note("How to read this chart","Shows which personal data types are most frequently involved. Special category data (health, biometric etc.) carries UK GDPR Article 9 obligations.")

# --- TAB 3 — IMPACT & SEVERITY -------------------------------------------------
with tabs[3]:
    st.markdown("### Breach Impact & Severity Analysis")
    if not filtered.empty:
        band_order=["1 to 9","10 to 99","100 to 1k","1k to 10k","10k to 100k","Over 100k"]
        imp=filtered.groupby(["No_Data_Subjects_Affected","Incident_Category"]).size().reset_index(name="Reports")
        fig_imp=px.bar(imp,x="No_Data_Subjects_Affected",y="Reports",color="Incident_Category",barmode="group",
                       color_discrete_map={"Cyber":C_CYBER,"Non Cyber":C_NONCYBER},category_orders={"No_Data_Subjects_Affected":band_order},template=TPL,height=380)
        fig_imp.update_layout(xaxis_title="People affected (band)",legend_title="Breach Category")
        st.plotly_chart(fig_imp,use_container_width=True)
        note("How to read this chart","X-axis = ordinal bands of people affected. Cyber breaches (red) tend to appear more in larger impact bands as they often target databases.")
        st.markdown("---")
        top12=filtered["Sector"].value_counts().head(12).index.tolist()
        fig_sv=px.box(filtered[filtered["Sector"].isin(top12)],x="Sector",y="Severity_Score",color="Incident_Category",
                      color_discrete_map={"Cyber":C_CYBER,"Non Cyber":C_NONCYBER},template=TPL,height=420,labels={"Severity_Score":"Severity Score (0-11)"})
        fig_sv.update_layout(xaxis_tickangle=-35,legend_title="Breach Category")
        st.plotly_chart(fig_sv,use_container_width=True)
        note("How to read this box plot","Box = interquartile range (25th–75th percentile). Line = median. Dots beyond whiskers = outlier breaches. Wider/higher boxes = more variable, more severe breaches.")
        st.markdown("---")
        all_yrs=sorted(df_full["Year"].dropna().unique().tolist())
        midpt=all_yrs[len(all_yrs)//2]
        p1,p2=st.columns(2)
        with p1: period_a=st.multiselect("Period A",all_yrs,default=[y for y in all_yrs if y<midpt],key="pa")
        with p2: period_b=st.multiselect("Period B",all_yrs,default=[y for y in all_yrs if y>=midpt],key="pb")
        if period_a and period_b:
            cmp_col=st.selectbox("Compare by",["Incident_Category","Sector","Incident_Type","Data_Type"],
                                 format_func=lambda x:{"Incident_Category":"Breach Category","Sector":"Sector","Incident_Type":"Breach Type","Data_Type":"Data Category"}.get(x,x),key="cmp_metric")
            da=df_full[df_full["Year"].isin(period_a)]; db=df_full[df_full["Year"].isin(period_b)]
            ca=da[cmp_col].value_counts().reset_index(); ca.columns=[cmp_col,"Period A"]
            cb=db[cmp_col].value_counts().reset_index(); cb.columns=[cmp_col,"Period B"]
            mg=ca.merge(cb,on=cmp_col,how="outer").fillna(0).sort_values("Period A",ascending=False).head(15)
            melt=mg.melt(id_vars=cmp_col,var_name="Period",value_name="Reports")
            fig_c=px.bar(melt,x="Reports",y=cmp_col,color="Period",barmode="group",orientation="h",template=TPL,height=420)
            fig_c.update_layout(yaxis=dict(categoryorder="total ascending"))
            st.plotly_chart(fig_c,use_container_width=True)
            note("How to read this chart","Longer Period B bars = increase. Helps identify growing or declining breach types between periods.")

# --- TAB 4 — FEATURE INSIGHTS --------------------------------------------------
with tabs[4]:
    st.markdown("### Feature Engineering Insights")
    st.markdown("Patterns derived from engineered features — metrics constructed to reveal deeper analytical insights.")
    if not filtered.empty:
        top10s=filtered["Sector"].value_counts().head(10).index.tolist()
        hm_data=filtered[filtered["Sector"].isin(top10s)].groupby(["Sector","Year"])["Severity_Score"].mean().round(2).reset_index()
        hm_pivot=hm_data.pivot(index="Sector",columns="Year",values="Severity_Score")
        fig_hm=px.imshow(hm_pivot,color_continuous_scale="RdYlGn_r",labels=dict(color="Avg Severity"),aspect="auto",template=TPL,height=420,title="Avg Severity Score by Sector & Year")
        st.plotly_chart(fig_hm,use_container_width=True)
        note("How to read this heatmap","Darker red = more severe on average. Tracks which sectors are getting worse over time.")
        st.markdown("---")
        dual=filtered[(filtered["Is_Cyber"]==1)&(filtered["Is_Special_Category"]==1)]
        if len(dual)>0:
            dh=dual["Sector"].value_counts().head(12).reset_index(); dh.columns=["Sector","High-Risk Breaches"]
            fig_dh=px.bar(dh,x="High-Risk Breaches",y="Sector",orientation="h",color_discrete_sequence=[C_PURPLE],template=TPL,height=400,title="Cyber breaches involving special category data")
            fig_dh.update_layout(yaxis=dict(categoryorder="total ascending")); st.plotly_chart(fig_dh,use_container_width=True)
            note("How to read this chart","Sectors with the highest count of breaches that are both cyber-classified AND involve special category data — the highest regulatory exposure combination.")
        st.markdown("---")
        fig_is=px.histogram(filtered,x="Impact_Score",color="Incident_Category",barmode="overlay",
                            color_discrete_map={"Cyber":C_CYBER,"Non Cyber":C_NONCYBER},opacity=0.7,template=TPL,height=340,nbins=6,
                            labels={"Impact_Score":"Impact Score (1=1-9 people, 6=Over 100k)"})
        fig_is.update_layout(legend_title="Breach Category"); st.plotly_chart(fig_is,use_container_width=True)
        note("How to read this histogram","Score 1 = smallest breaches. Score 6 = largest (100k+). Cyber breaches tend to concentrate at higher scores.")
        st.markdown("---")
        dec_sev=filtered.groupby("Decision_Taken")["Severity_Score"].mean().round(2).reset_index().sort_values("Severity_Score",ascending=False)
        fig_ds=px.bar(dec_sev,x="Severity_Score",y="Decision_Taken",orientation="h",color="Severity_Score",
                      color_continuous_scale="RdYlGn_r",template=TPL,height=360,labels={"Severity_Score":"Avg Severity Score","Decision_Taken":"Regulatory Decision"})
        fig_ds.update_layout(coloraxis_showscale=False,yaxis=dict(categoryorder="total ascending")); st.plotly_chart(fig_ds,use_container_width=True)
        note("How to read this chart","If investigation-type decisions have higher severity scores, it confirms the ICO's enforcement activity correlates with breach severity.")
        st.markdown("---")
        top8s=filtered["Sector"].value_counts().head(8).index.tolist()
        tm_df=filtered[filtered["Sector"].isin(top8s)].groupby(["Sector","Incident_Category","Decision_Taken"]).size().reset_index(name="Reports")
        fig_tm=px.treemap(tm_df,path=["Sector","Incident_Category","Decision_Taken"],values="Reports",color="Reports",color_continuous_scale="Blues",template=TPL,height=480)
        fig_tm.update_layout(margin=dict(t=30,b=10)); st.plotly_chart(fig_tm,use_container_width=True)
        note("How to read this treemap","Area = volume of reports. Click a sector to drill down into its cyber/non-cyber split and regulatory outcomes.")

# --- TAB 5 — DATA QUALITY ------------------------------------------------------
with tabs[5]:
    st.markdown("### Data Quality & Completeness")
    miss=df_full.isnull().sum().reset_index(); miss.columns=["Field","Missing Values"]
    miss["Missing (%)"]=(miss["Missing Values"]/len(df_full)*100).round(1)
    st.dataframe(miss.sort_values("Missing Values",ascending=False),use_container_width=True)
    note("Why missing values matter","Missing fields can distort sector/type-level charts. Records without a Sector label are excluded from sector charts, which may undercount certain industries.")
    if not filtered.empty:
        yr_cnt=df_full["Year"].value_counts().sort_index().reset_index(); yr_cnt.columns=["Year","Breach Reports"]
        fig_yr=px.bar(yr_cnt,x="Year",y="Breach Reports",template=TPL,color_discrete_sequence=[C_NEUTRAL],title="Total breach reports received by the ICO — by year")
        st.plotly_chart(fig_yr,use_container_width=True)
        note("About the 2025 data","This version uses full Q1–Q4 2025 data. All years are fully comparable.")
        pivot=filtered.groupby(["Decision_Taken","Incident_Category"]).size().reset_index(name="Reports")
        fig_pv=px.density_heatmap(pivot,x="Incident_Category",y="Decision_Taken",z="Reports",color_continuous_scale="Blues",template=TPL,height=420)
        st.plotly_chart(fig_pv,use_container_width=True)
        note("How to read this heatmap","Darker = more reports in that combination. Cyber breaches in investigation cells suggests they attract more regulatory scrutiny.")

# TAB 6 - PREDICTIVE MODEL
with tabs[6]:
    st.markdown("### Predicting Breach Category: Cyber vs Non-Cyber")
    @st.cache_resource(show_spinner=False)
    def get_model_results():
        return train_models(df_full)
    model_results = get_model_results()
    best_name  = max(model_results,key=lambda k:model_results[k]["auc"])
    m_ch=st.selectbox("Select model",list(model_results.keys()),key="mc")
    res=model_results[m_ch]; rep=res["report"]
    mc1,mc2,mc3,mc4=st.columns(4)
    mc1.metric("Accuracy",str(round(rep["accuracy"],3)))
    mc2.metric("Precision — Cyber",str(round(rep.get("1",{}).get("precision",0),3)))
    mc3.metric("Recall — Cyber",str(round(rep.get("1",{}).get("recall",0),3)))
    mc4.metric("ROC-AUC",str(round(res["auc"],3)))
    note("What these metrics mean","**Accuracy** = overall correctness. **Precision** = of predicted cyber, how many truly are. **Recall** = of real cyber breaches, how many the model caught. **ROC-AUC** above 0.8 is strong for this classification task.")
    st.markdown("---")
    rc1,rc2=st.columns(2)
    with rc1:
        fig_roc=go.Figure()
        fig_roc.add_trace(go.Scatter(x=res["fpr"],y=res["tpr"],mode="lines",name=m_ch+" (AUC="+str(round(res["auc"],3))+")",line=dict(color=C_CYBER,width=2)))
        fig_roc.add_shape(type="line",x0=0,y0=0,x1=1,y1=1,line=dict(dash="dash",color="grey"))
        fig_roc.update_layout(xaxis_title="False Positive Rate",yaxis_title="True Positive Rate",template=TPL,height=360)
        st.plotly_chart(fig_roc,use_container_width=True)
        note("How to read the ROC Curve","Curve hugging top-left = better model. Dashed diagonal = random guessing baseline.")
    with rc2:
        fig_cm=px.imshow(res["cm"],labels=dict(x="Predicted",y="Actual",color="Reports"),x=["Non-Cyber","Cyber"],y=["Non-Cyber","Cyber"],text_auto=True,color_continuous_scale="Blues",template=TPL,height=360)
        st.plotly_chart(fig_cm,use_container_width=True)
        note("How to read the Confusion Matrix","Top-left = correct Non-Cyber. Bottom-right = correct Cyber. Off-diagonal = errors.")
    if m_ch=="Random Forest" and res.get("fi") is not None:
        fig_fi=px.bar(res["fi"].sort_values("Importance"),x="Importance",y="Feature",orientation="h",template=TPL,height=520,color="Importance",color_continuous_scale="Blues")
        fig_fi.update_layout(yaxis=dict(categoryorder="total ascending"),coloraxis_showscale=False); st.plotly_chart(fig_fi,use_container_width=True)
        note("How to read Feature Importance","Higher bars = that feature contributes more to predicting cyber vs non-cyber. Engineered features (Impact_Score, Is_Special_Category) appearing high confirm they add predictive value.")
    comp=[]
    for k,v in model_results.items():
        comp.append({"Model":k,"Accuracy":round(v["report"]["accuracy"],3),"Precision (Cyber)":round(v["report"].get("1",{}).get("precision",0),3),"Recall (Cyber)":round(v["report"].get("1",{}).get("recall",0),3),"F1 (Cyber)":round(v["report"].get("1",{}).get("f1-score",0),3),"ROC-AUC":round(v["auc"],3),"Best?":"✅" if k==best_name else ""})
    st.dataframe(pd.DataFrame(comp),use_container_width=True,hide_index=True)

# TAB 7 - RISK PREDICTOR
with tabs[7]:
    st.markdown("### Cyber Breach Risk Estimator")
    st.markdown("Model: **"+best_name+"** (ROC-AUC = "+str(round(model_results[best_name]["auc"],3))+")")
    st.warning("**Important:** For exploratory use only. Not for operational breach reporting. Visit [ico.org.uk/report-a-breach](https://ico.org.uk/for-organisations/report-a-breach/) for guidance.")
    band_opts=["1 to 9","10 to 99","100 to 1k","1k to 10k","10k to 100k","Over 100k"]
    pA,pB,pC=st.columns(3)
    with pA:
        p_s=st.selectbox("Sector",sorted(df_full["Sector"].dropna().unique().tolist()),key="p_s")
        p_dst=st.selectbox("Who was affected",sorted(df_full["Data_Subject_Type"].dropna().unique().tolist()),key="p_dst")
    with pB:
        p_dt=st.selectbox("Data category",sorted(df_full["Data_Type"].dropna().unique().tolist()),key="p_dt")
        p_it=st.selectbox("How did the breach occur",sorted(df_full["Incident_Type"].dropna().unique().tolist()),key="p_it")
    with pC:
        p_band=st.selectbox("People affected",band_opts,key="p_band")
        p_time=st.selectbox("Time to report",sorted(df_full["Time_Taken_to_Report"].dropna().unique().tolist()),key="p_time")
        p_year=st.selectbox("Year",sorted(df_full["Year"].dropna().unique().tolist()),key="p_year")
    if st.button("Estimate cyber breach probability",use_container_width=True):
        bs={"1 to 9":1,"10 to 99":2,"100 to 1k":3,"1k to 10k":4,"10k to 100k":5,"Over 100k":6}
        sc_kw=["health","racial","ethnic","biometric","genetic","sexual","religion","political","criminal"]
        is_sc=int(any(k in p_dt.lower() for k in sc_kw))
        imp_s=bs.get(p_band,1)
        w72=1 if any(t in p_time.lower() for t in ["0 to 24","24 to 48","48 to 72"]) else 0
        X_new=pd.DataFrame([{"Sector":p_s,"Data_Subject_Type":p_dst,"Data_Type":p_dt,"Incident_Type":p_it,"No_Data_Subjects_Affected":p_band,"Time_Taken_to_Report":p_time,"Year":p_year,"Is_Special_Category":is_sc,"Impact_Score":imp_s,"Within_72hrs":w72}])
        proba=best_pipe.predict_proba(X_new)[0,1]
        label="Cyber" if proba>=0.5 else "Non-Cyber"
        sev_s=3*int(proba>=0.5)+imp_s+is_sc*2
        pct_s=str(round(proba*100,1))+"%"
        g_col,t_col=st.columns([1,2])
        with g_col:
            fig_g=go.Figure(go.Indicator(mode="gauge+number",value=round(proba*100,1),number={"suffix":"%"},title={"text":"Cyber Probability"},gauge={"axis":{"range":[0,100]},"bar":{"color":C_CYBER if proba>=0.5 else C_NONCYBER},"steps":[{"range":[0,40],"color":"#dcfce7"},{"range":[40,70],"color":"#fef9c3"},{"range":[70,100],"color":"#fee2e2"}],"threshold":{"line":{"color":"black","width":3},"value":50}}))
            fig_g.update_layout(height=300,margin=dict(t=60,b=20)); st.plotly_chart(fig_g,use_container_width=True)
        with t_col:
            msg="Predicted: **"+label+"**  \nCyber probability: **"+pct_s+"**  \nSeverity Score: **"+str(sev_s)+" / 11**  \nSpecial category data: **"+("Yes" if is_sc else "No")+"**  \nModel: "+best_name
            (st.error if label=="Cyber" else st.success)(msg)

# TAB 8 - KEY INSIGHTS
with tabs[8]:
    st.markdown("### Key Findings")
    if filtered.empty:
        st.warning("No data available.")
    else:
        for ins in insights(filtered):
            st.info(ins)
        st.markdown("---")
        dec=filtered.groupby(["Decision_Taken","Incident_Category"]).size().reset_index(name="Reports")
        fig_dec=px.bar(dec,x="Decision_Taken",y="Reports",color="Incident_Category",barmode="group",
                       color_discrete_map={"Cyber":C_CYBER,"Non Cyber":C_NONCYBER},template=TPL,height=380)
        st.plotly_chart(fig_dec,use_container_width=True)
        note("How to read this chart","If cyber breaches more frequently result in 'Investigation Pursued', this aligns with ICO guidance that cyber breaches carry higher risk of harm to individuals.")
        st.markdown("---")
        st.markdown("#### Download filtered data")
        dl1,dl2=st.columns(2)
        with dl1:
            csv_bytes=filtered.to_csv(index=False).encode("utf-8")
            st.download_button("⬇️  Download as CSV",csv_bytes,"ico_breach_data_filtered.csv","text/csv",use_container_width=True,help="Plain CSV with all columns including engineered features.")
        with dl2:
            try:
                excel_bytes=build_excel_export(filtered)
                st.download_button("📊  Download as Excel (styled workbook)",excel_bytes,"ICO_Breach_Data_Export.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True,help="5-sheet Excel: Data · Dictionary · Methods · Summary Stats · Overview.")
            except Exception as e:
                st.error("Excel export failed: "+str(e))

# TAB 9 - ABOUT THE DATA
with tabs[9]:
    st.markdown("### About This Dashboard & the Underlying Data")
    st.markdown("""
#### Data Source
Built on the **ICO Data Security Incident Trends** dataset published by the [UK Information Commissioner's Office (ICO)](https://ico.org.uk).

> *"We publish this information to help organisations understand what to look out for and take appropriate action."* — ICO

---

#### What the data contains
Each row = one personal data breach self-reported to the ICO.

> *"A breach of security leading to the accidental or unlawful destruction, loss, alteration, unauthorised disclosure of, or access to, personal data."* — UK GDPR Article 4(12)

| Field | ICO Definition |
|---|---|
| **Breach Category** | Cyber (malicious, third-party) or Non-Cyber (human error, physical loss) |
| **Breach Type** | How the breach occurred |
| **Data Subject Type** | Who was affected: customers, employees, patients, etc. |
| **Data Category** | Type of personal data, including special category data |
| **Regulatory Decision** | ICO response: No Further Action, Investigation Pursued, etc. |
| **People Affected** | Estimated band of individuals impacted |
| **Sector** | Organisation type |
| **Time to Report** | Hours from discovery to notification |

---

#### Engineered Features

| Feature | Construction | Range |
|---|---|---|
| **Severity Score** | 3pts (cyber) + 1–6pts (impact) + 2pts (special category) | 0–11 |
| **Is Special Category** | 1 if Data Category contains Art.9 keywords | 0/1 |
| **Impact Score** | Ordinal encoding of people-affected band | 1–6 |
| **Sector Risk Tier** | High/Medium/Low based on cyber rate tertiles | 3 levels |
| **Within 72hrs** | 1 if reported within Art.33 timeframe | 0/1 |

---

#### Limitations
- Covers **self-reported breaches only** — unreported incidents are not captured.
- **Sector labels** are not always consistent in historic records.
- People-affected figures are **estimates at time of notification**.
- Engineered features are **analytical constructs** — not ICO classifications.

---

#### Further reading
- [ICO Data Security Incident Trends](https://ico.org.uk/action-weve-taken/complaints-and-concerns-data-sets/data-security-incident-trends/)
- [ICO Glossary of Terms](https://ico.org.uk/action-weve-taken/complaints-and-concerns-data-sets/data-security-incident-trends/glossary-of-terms/)
- [UK GDPR Article 33 — Breach Notification](https://ico.org.uk/for-organisations/report-a-breach/personal-data-breach/)
- [ICO: Responding to a Cybersecurity Incident](https://ico.org.uk/media2/migrated/2614816/responding-to-a-cybersecurity-incident.pdf)
""")

st.markdown("---")
st.caption("Data: UK Information Commissioner's Office — ico.org.uk | Built with Streamlit and Plotly")